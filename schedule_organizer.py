#!/usr/bin/env python3
import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def organize_schedule_data():
    """
    Organizes the consolidated schedule data into a structured Excel workbook
    with multiple sheets for better organization and analysis.
    """
    # Check if consolidated schedule data exists
    if not os.path.exists('consolidated_schedules.csv'):
        print("Consolidated schedule data not found. Please run daily_schedule_processor.py first.")
        return False
    
    print("Reading consolidated schedule data...")
    # Read the consolidated schedule data
    data = pd.read_csv('consolidated_schedules.csv')
    
    # Basic data cleaning
    # Convert date columns to datetime if they exist
    date_columns = [col for col in data.columns if 'date' in col.lower()]
    for col in date_columns:
        try:
            data[col] = pd.to_datetime(data[col], errors='coerce')
        except:
            pass  # Skip if conversion fails
    
    # Create Excel workbook
    print("Creating Excel workbook...")
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ---- Sheet 1: All Jobs ----
    ws_all = wb.create_sheet("All Jobs")
    
    # Add data to sheet
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            # Format header row
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_all.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws_all.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width at 50
    
    # ---- Sheet 2: Jobs by Date ----
    # Sort data by date if date column exists
    if date_columns:
        primary_date_col = date_columns[0]  # Use first date column
        date_data = data.sort_values(by=primary_date_col, na_position='last')
        
        ws_date = wb.create_sheet("Jobs by Date")
        
        # Add data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(date_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_date.cell(row=r_idx, column=c_idx, value=value)
                # Format header row
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws_date.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_date.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width at 50
    
    # ---- Sheet 3: Jobs by Client ----
    # Identify client column (look for client, customer, name, etc.)
    client_columns = [col for col in data.columns if any(term in col.lower() 
                     for term in ['client', 'customer', 'name', 'company'])]
    
    if client_columns:
        client_col = client_columns[0]  # Use first client column
        
        # Group by client
        client_groups = data.groupby(client_col)
        
        ws_client = wb.create_sheet("Jobs by Client")
        
        # Create header
        headers = list(data.columns)
        for c_idx, header in enumerate(headers, 1):
            cell = ws_client.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Add data for each client with a separator row
        row_idx = 2
        for client, group in client_groups:
            # Add client as a header
            ws_client.cell(row=row_idx, column=1, value=f"Client: {client}")
            ws_client.cell(row=row_idx, column=1).font = Font(bold=True)
            ws_client.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            row_idx += 1
            
            # Add group data
            for _, row in group.iterrows():
                for c_idx, value in enumerate(row, 1):
                    ws_client.cell(row=row_idx, column=c_idx, value=value)
                row_idx += 1
            
            # Add separator
            row_idx += 1
        
        # Auto-adjust column widths
        for column in ws_client.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_client.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Cap width at 50
    
    # ---- Sheet 4: Summary Statistics ----
    ws_summary = wb.create_sheet("Summary Statistics")
    
    # Add basic statistics
    ws_summary.cell(row=1, column=1, value="Summary Statistics").font = Font(bold=True, size=14)
    
    row_idx = 3
    ws_summary.cell(row=row_idx, column=1, value="Total Jobs:").font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=2, value=len(data))
    row_idx += 1
    
    # Jobs per date if date column exists
    if date_columns:
        primary_date_col = date_columns[0]
        
        # Count jobs per date
        jobs_per_date = data.groupby(pd.Grouper(key=primary_date_col, freq='D')).size()
        
        if not jobs_per_date.empty:
            ws_summary.cell(row=row_idx, column=1, value="Average Jobs per Day:").font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=jobs_per_date.mean())
            row_idx += 1
            
            row_idx += 1
            ws_summary.cell(row=row_idx, column=1, value="Top 10 Busiest Days:").font = Font(bold=True)
            row_idx += 1
            
            ws_summary.cell(row=row_idx, column=1, value="Date").font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value="Number of Jobs").font = Font(bold=True)
            row_idx += 1
            
            for date, count in jobs_per_date.nlargest(10).items():
                ws_summary.cell(row=row_idx, column=1, value=date)
                ws_summary.cell(row=row_idx, column=2, value=count)
                row_idx += 1
    
    # Jobs per client if client column exists
    if client_columns:
        client_col = client_columns[0]
        
        row_idx += 2
        ws_summary.cell(row=row_idx, column=1, value="Jobs per Client:").font = Font(bold=True)
        row_idx += 1
        
        ws_summary.cell(row=row_idx, column=1, value="Client").font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value="Number of Jobs").font = Font(bold=True)
        row_idx += 1
        
        # Count jobs per client
        client_counts = data[client_col].value_counts()
        
        for client, count in client_counts.items():
            if pd.notna(client) and client:  # Skip empty or NaN clients
                ws_summary.cell(row=row_idx, column=1, value=client)
                ws_summary.cell(row=row_idx, column=2, value=count)
                row_idx += 1
    
    # Auto-adjust column widths for summary sheet
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws_summary.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # Save workbook
    output_file = 'organized_schedules.xlsx'
    print(f"Saving organized data to {output_file}...")
    wb.save(output_file)
    print(f"Successfully created {output_file} with organized schedule data!")
    return True

if __name__ == "__main__":
    organize_schedule_data() 